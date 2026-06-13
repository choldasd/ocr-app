"""
export.py
─────────
Service for exporting batch OCR results in multiple formats:
  JSON, CSV, Excel (.xlsx), PDF.
"""

import io
import csv
import json
import logging
from typing import List

logger = logging.getLogger(__name__)


def build_export_rows(jobs) -> List[dict]:
    """
    Converts a list of Job ORM objects into a flat list of row dicts.
    Each page becomes a separate row.
    """
    rows = []
    for job in jobs:
        if job.ocr_result and isinstance(job.ocr_result, dict):
            pages = job.ocr_result.get("pages", [])
            for page in pages:
                rows.append({
                    "file_name": job.file_name,
                    "job_status": job.status,
                    "page": page.get("page", 1),
                    "engine": page.get("engine", ""),
                    "confidence": page.get("confidence", 0.0),
                    "text": page.get("text", ""),
                    "error_message": job.error_message or "",
                })
        else:
            rows.append({
                "file_name": job.file_name,
                "job_status": job.status,
                "page": 1,
                "engine": "",
                "confidence": 0.0,
                "text": job.plain_text or "",
                "error_message": job.error_message or "",
            })
    return rows


# ─── JSON export ──────────────────────────────────────────────────────────────

def export_json(batch, jobs) -> bytes:
    payload = {
        "batch_uuid": batch.batch_uuid,
        "status": batch.status,
        "total_files": batch.total_files,
        "completed_files": batch.completed_files,
        "failed_files": batch.failed_files,
        "jobs": [
            {
                "file_name": j.file_name,
                "status": j.status,
                "retry_count": j.retry_count,
                "error_message": j.error_message,
                "ocr_result": j.ocr_result,
                "plain_text": j.plain_text,
            }
            for j in jobs
        ],
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


# ─── CSV export ───────────────────────────────────────────────────────────────

def export_csv(batch, jobs) -> bytes:
    rows = build_export_rows(jobs)
    output = io.StringIO()
    fieldnames = ["file_name", "job_status", "page", "engine", "confidence", "text", "error_message"]
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue().encode("utf-8")


# ─── Excel export ─────────────────────────────────────────────────────────────

def export_excel(batch, jobs) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl is not installed.")

    rows = build_export_rows(jobs)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OCR Results"

    headers = ["File Name", "Status", "Page", "Engine", "Confidence", "Text", "Error"]
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    col_widths = [35, 14, 8, 14, 12, 80, 30]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    key_map = ["file_name", "job_status", "page", "engine", "confidence", "text", "error_message"]
    for row_data in rows:
        ws.append([row_data.get(k, "") for k in key_map])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── PDF export ───────────────────────────────────────────────────────────────

def export_pdf(batch, jobs) -> bytes:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        )
    except ImportError:
        raise RuntimeError("reportlab is not installed.")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"],
                                 fontSize=18, textColor=colors.HexColor("#1e40af"))
    heading_style = ParagraphStyle("heading", parent=styles["Heading2"],
                                   fontSize=12, textColor=colors.HexColor("#1d4ed8"))
    body_style = ParagraphStyle("body", parent=styles["Normal"],
                                fontSize=9, leading=13, wordWrap="CJK")

    story = []
    story.append(Paragraph(f"Batch OCR Export — {batch.batch_uuid}", title_style))
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph(
        f"Status: {batch.status}  |  Total: {batch.total_files}  |  "
        f"Completed: {batch.completed_files}  |  Failed: {batch.failed_files}",
        styles["Normal"]
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#dbeafe")))
    story.append(Spacer(1, 0.4*cm))

    for job in jobs:
        story.append(Paragraph(f"📄 {job.file_name}  [{job.status}]", heading_style))
        if job.error_message:
            story.append(Paragraph(f"Error: {job.error_message}", body_style))
        if job.ocr_result and isinstance(job.ocr_result, dict):
            for page in job.ocr_result.get("pages", []):
                story.append(Paragraph(
                    f"<b>Page {page.get('page', '?')} — {page.get('engine', '')} "
                    f"(conf: {page.get('confidence', 0.0):.2f})</b>",
                    body_style
                ))
                text = (page.get("text") or "").replace("\n", "<br/>")
                story.append(Paragraph(text or "(no text extracted)", body_style))
        elif job.plain_text:
            story.append(Paragraph(job.plain_text.replace("\n", "<br/>"), body_style))
        story.append(Spacer(1, 0.3*cm))

    doc.build(story)
    return buf.getvalue()
